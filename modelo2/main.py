from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
import os


def create_certificate(nome_aluno ,cpf ,curso, carga_horaria, nome_coordenador, titulo_coordenador, nome_diretor, titulo_diretor , data_evento_completo):
    os.makedirs('./todos', exist_ok=True)
    page_size = (3509, 2481)

    certificado_base = "template.pdf"
    certificado_temp = f"./todos/temp_certificado_{nome_aluno.replace(' ', '_')}.pdf"
    certificado_final = f"./todos/{nome_aluno.replace(' ', '_')}.pdf"

    c = canvas.Canvas(certificado_temp, pagesize=page_size)
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 40)
    c.drawCentredString(550, 1240, f"Certificamos que")

    c.setFont("Helvetica", 18)
    c.drawString(50, 400, f"Certificamos que {nome_aluno}    CPF {cpf}")
    c.setFont("Helvetica", 16)
    c.drawString(190, 400, f"{len(nome_aluno) * '_'}")
    c.setFont("Helvetica", 18)
    c.drawString(50, 380, f"participou no evento promovido pela Faculdade de Tecnologia de Rio Claro")
    c.setFillColor(colors.red)
    c.setFont("Helvetica", 25)
    c.drawString(150, 340, f"{curso}")

    c.setFont("Helvetica", 18)
    c.setFillColor(colors.black)
    c.drawString(50, 300, f"realizada no dia {data_evento_completo}, com carga horária de                 horas.")
    
    c.setFont("Helvetica-Bold", 18)
    c.drawString(570, 300, f"{carga_horaria} ")

    c.setFont("Helvetica", 18)
    c.drawString(550, 220, f"Rio Claro, {data_evento_completo}")

    c.setFont("Helvetica", 13)
    c.drawString(100, 120, f"{nome_coordenador}")
    c.drawString(100, 100, f"{titulo_coordenador}")
    c.drawString(500, 120, f"{nome_diretor}")
    c.drawString(500, 100, f"{titulo_diretor}")
    c.save()

    # Mescla com o template base
    background = PdfReader(certificado_base)
    overlay = PdfReader(certificado_temp)
    output = PdfWriter()

    background_page = background.pages[0]
    overlay_page = overlay.pages[0]
    background_page.merge_page(overlay_page)
    output.add_page(background_page)

    with open(certificado_final, "wb") as out_file:
        output.write(out_file)

    os.remove(certificado_temp)

    print(f"✅ Certificado gerado: {certificado_final}")


# Dados Evento
planilha_alunos = load_workbook('./alunos.xlsx')
evento_sheet = planilha_alunos['Evento']
curso = evento_sheet['B1'].value
carga_horaria = evento_sheet['B2'].value
nome_coordenador = evento_sheet['B3'].value
titulo_coordenador = evento_sheet['B4'].value
nome_diretor = evento_sheet['B5'].value
titulo_diretor = evento_sheet['B6'].value
data_evento_completo = evento_sheet['B7'].value

pagina_certificados = planilha_alunos['certificados']
# Gera certificados
for linha in pagina_certificados.iter_rows(min_row=2, values_only=True):
    nome_aluno, cpf = linha
    create_certificate(nome_aluno ,cpf ,curso, carga_horaria, nome_coordenador, titulo_coordenador, nome_diretor, titulo_diretor , data_evento_completo)
