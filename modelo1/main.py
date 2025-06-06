from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Spacer, Image, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from datetime import datetime
from openpyxl import load_workbook
import os


# Função para criar o certificado em PDF
def create_certificate(nome_aluno, cpf_aluno, curso, carga_horaria, responsavel):
    # Cria diretório 'todos' caso não exista
    os.makedirs('./todos', exist_ok=True)

    # Caminho para o PDF
    filename = f'./todos/certificado_{nome_aluno.replace(" ", "_")}.pdf'

    # Cria o documento
    doc = SimpleDocTemplate(filename, pagesize=letter)

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='Title',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.black,
        alignment=1,
        spaceAfter=20
    )

    body_style = ParagraphStyle(
        name='BodyText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        textColor=colors.black,
        alignment=1,
        spaceAfter=20
    )

    # Logo (opcional)
    content = []
    logo_path = "./logo.png"
    if os.path.exists(logo_path):
        im = Image(logo_path, width=160, height=90)
        content.append(im)

    # Texto principal
    today = datetime.now().strftime("%d/%m/%Y")
    content.extend([
        Spacer(1, 12),
        Paragraph(f'<b>CERTIFICADO</b>', title_style),
        Spacer(1, 20),
        Paragraph(
            f"Este certificado atesta que <b>{nome_aluno}</b>, portador do CPF <b>{cpf_aluno}</b>, concluiu com êxito o curso de <b>{curso}</b>, totalizando <b>{int(carga_horaria)}</b> horas de dedicação e aprendizado.",
            body_style),
        Spacer(1, 20),
        Paragraph(
            f"Dado o exposto, concedemos este certificado como reconhecimento oficial de que <b>{nome_aluno}</b> completou com sucesso o curso, atestando sua competência e comprometimento.",
            body_style),
        Spacer(1, 20),
        Paragraph(f"<b>Data de Emissão:</b> {today}", body_style),
        Spacer(1, 40),
        Paragraph(f"<b>Responsável:</b> <i>{responsavel}</i>", body_style),
    ])

    # Gera o PDF
    doc.build(content)
    print(f"Certificado gerado: {filename}")


# Carrega a planilha de alunos
planilha_alunos = load_workbook('./alunos.xlsx')

# Captura dados da planilha "Evento"
evento_sheet = planilha_alunos['Evento']
curso = evento_sheet['B1'].value
carga_horaria = evento_sheet['B2'].value
responsavel = evento_sheet['B3'].value

# Acessa a planilha "certificados"
pagina_certificados = planilha_alunos['certificados']

# Itera pelas linhas (ignorando o cabeçalho)
for linha in pagina_certificados.iter_rows(min_row=2, values_only=True):
    nome_aluno, cpf_aluno = linha
    create_certificate(nome_aluno, cpf_aluno, curso, carga_horaria, responsavel)

